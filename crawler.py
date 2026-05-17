import os
import time
import requests
import pandas as pd
from io import BytesIO
from fractions import Fraction
from PIL import Image
import imagehash
from openpyxl import load_workbook
from openpyxl.styles import Alignment

API_KEY = "a085ca875a968a2a3296d7b1bf729808"

BASE_DIR = "photo_data"

TARGET_COUNT = 6000
PER_PAGE = 100
MAX_PAGES_PER_QUERY = 50
NO_PROGRESS_LIMIT = 6
REQUEST_TIMEOUT = 20

# 중복/품질 설정
PHASH_DUP_THRESHOLD = 5
MAX_IMAGES_PER_OWNER = 30
MIN_WIDTH = 500
MIN_HEIGHT = 500
MIN_COLOR_COUNT = 50

os.makedirs(BASE_DIR, exist_ok=True)

# =========================================================
# 카테고리별 검색어/키워드
# =========================================================
CATEGORY_CONFIG = {
    "food": {
        "queries": [
            "korean food photography", "japanese food photography", "chinese food photography",
            "italian food photography", "western food dish", "street food photography",
            "restaurant food plating", "fine dining food", "home cooked meal",
            "breakfast food photography", "lunch meal photography", "dinner dish photography",
            "dessert photography", "cake dessert photography", "bakery bread photography",
            "pasta dish photography", "pizza food photography", "burger food photography",
            "seafood dish photography", "vegetarian food photography", "salad food photography",
            "noodle dish photography", "soup food photography", "rice bowl food photography"
        ],
        "positive_keywords": [
            "food", "dish", "meal", "dessert", "cake", "bread", "bakery",
            "pasta", "pizza", "burger", "seafood", "salad", "soup",
            "noodle", "rice", "breakfast", "lunch", "dinner", "restaurant",
            "plating", "cuisine", "korean", "japanese", "chinese", "italian",
            "streetfood", "vegetarian"
        ],
        "negative_keywords": [
            "landscape", "mountain", "sky", "forest", "portrait", "selfie",
            "wedding", "car", "street scene", "building", "menu", "kitchen",
            "market", "sign", "logo", "single egg", "raw egg",
            "ingredient", "grocery", "package", "packaging"
        ]
    },

    "landscape": {
        "queries": [
            "mountain landscape photography", "sea landscape photography", "ocean view landscape",
            "sky landscape photography", "forest landscape photography", "lake landscape photography",
            "river landscape photography", "waterfall landscape photography", "valley landscape photography",
            "desert landscape photography", "snow mountain landscape", "sunset landscape photography",
            "sunrise landscape photography", "beach landscape photography", "field landscape photography",
            "countryside landscape", "natural scenery photography", "cliff landscape photography",
            "island landscape photography", "autumn landscape photography", "winter landscape photography"
        ],
        "positive_keywords": [
            "landscape", "scenery", "mountain", "sea", "ocean", "sky",
            "forest", "lake", "river", "waterfall", "valley", "desert",
            "nature", "sunset", "sunrise", "beach", "field", "countryside",
            "cliff", "island", "snow", "autumn", "winter"
        ],
        "negative_keywords": [
            "portrait", "selfie", "food", "dish", "meal", "wedding",
            "car", "motorcycle", "building interior", "menu", "product",
            "indoor", "room", "restaurant", "person closeup", "face",
            "street portrait"
        ]
    },

    "night": {
        "queries": [
            "night city photography", "night street photography", "nightscape photography",
            "city night lights", "low light photography", "neon night photography",
            "night skyline photography", "night road photography", "night market photography",
            "night bridge photography", "night architecture photography", "night urban photography",
            "evening city photography", "long exposure night photography", "dark street photography",
            "night traffic lights", "night harbor photography"
        ],
        "positive_keywords": [
            "night", "nightscape", "city", "lights", "street", "dark",
            "low light", "midnight", "neon", "evening", "skyline",
            "urban", "bridge", "traffic", "harbor", "long exposure"
        ],
        "negative_keywords": [
            "food", "dish", "meal", "portrait studio", "daylight",
            "sunny", "morning", "afternoon", "beach day", "wedding",
            "product", "menu", "indoor food"
        ]
    },

    "portrait": {
        "queries": [
            "portrait photography", "face portrait photography", "person portrait photography",
            "outdoor portrait photography", "close up portrait photography", "natural light portrait",
            "studio portrait photography", "female portrait photography", "male portrait photography",
            "child portrait photography", "fashion portrait photography", "street portrait photography",
            "black and white portrait", "environmental portrait", "headshot photography",
            "model portrait photography", "candid portrait photography"
        ],
        "positive_keywords": [
            "portrait", "face", "person", "people", "woman", "man",
            "girl", "boy", "human", "model", "headshot", "fashion",
            "candid", "studio", "outdoor", "environmental"
        ],
        "negative_keywords": [
            "food", "dish", "meal", "landscape", "mountain", "ocean",
            "forest", "building", "car", "animal", "dog", "cat",
            "statue", "doll", "painting", "poster"
        ]
    },

    "contrast": {
        "queries": [
            "high contrast photography", "dramatic lighting photography",
            "light and shadow photography", "shadow photography",
            "contrast lighting photography", "strong shadows photography",
            "silhouette photography", "black and white high contrast",
            "moody light photography", "hard light photography",
            "chiaroscuro photography", "low key photography",
            "dramatic shadow portrait", "geometric shadow photography",
            "sunlight shadow photography", "dark contrast photography"
        ],
        "positive_keywords": [
            "contrast", "dramatic", "shadow", "light", "lighting",
            "silhouette", "high contrast", "moody", "hard light",
            "chiaroscuro", "low key", "dark", "black and white",
            "bw", "monochrome"
        ],
        "negative_keywords": [
            "menu", "product", "food menu", "building plan", "screenshot",
            "text", "poster", "graphic", "logo", "diagram"
        ]
    }
}

# =========================================================
# Flickr API
# =========================================================
def flickr_call(method: str, params: dict):
    url = "https://www.flickr.com/services/rest/"
    base_params = {
        "method": method,
        "api_key": API_KEY,
        "format": "json",
        "nojsoncallback": 1
    }
    base_params.update(params)

    response = requests.get(url, params=base_params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()

    if data.get("stat") != "ok":
        raise RuntimeError(f"Flickr API 오류: {data}")

    return data


def search_photos(query: str, page: int):
    data = flickr_call(
        "flickr.photos.search",
        {
            "text": query,
            "page": page,
            "per_page": PER_PAGE,
            "sort": "relevance",
            "safe_search": 1,
            "content_types": 0,
            "media": "photos",
            "extras": "url_l,url_c,url_z,url_m,tags,title,owner"
        }
    )
    return data["photos"]["photo"]


def choose_image_url(photo: dict):
    for key in ["url_l", "url_c", "url_z", "url_m"]:
        if photo.get(key):
            return photo[key]
    return None


def is_relevant_photo(photo: dict, positive_keywords: list, negative_keywords: list):
    title = str(photo.get("title", "")).lower()
    tags = str(photo.get("tags", "")).lower()
    combined = f"{title} {tags}"

    for bad in negative_keywords:
        if bad.lower() in combined:
            return False

    for good in positive_keywords:
        if good.lower() in combined:
            return True

    return False


# =========================================================
# EXIF 처리
# =========================================================
def get_exif(photo_id: str):
    try:
        data = flickr_call("flickr.photos.getExif", {"photo_id": photo_id})
        exif_list = data.get("photo", {}).get("exif", [])
    except Exception:
        return {}

    exif_map = {}

    for item in exif_list:
        tag = item.get("tag", "")
        label = item.get("label", "")
        raw = item.get("raw", {})

        if isinstance(raw, dict):
            raw_value = raw.get("_content", "")
        else:
            raw_value = str(raw)

        if tag:
            exif_map[tag] = raw_value
        if label:
            exif_map[label] = raw_value

    return exif_map


def extract_required_fields(exif_map: dict):
    iso = (
        exif_map.get("ISO")
        or exif_map.get("ISO Speed")
        or exif_map.get("ISOSpeedRatings")
        or ""
    )

    shutter_speed = (
        exif_map.get("ExposureTime")
        or exif_map.get("Exposure")
        or exif_map.get("Shutter Speed")
        or ""
    )

    focal_length = (
        exif_map.get("FocalLength")
        or exif_map.get("Focal Length")
        or ""
    )

    exposure_bias = (
        exif_map.get("ExposureBiasValue")
        or exif_map.get("Exposure Bias")
        or exif_map.get("Exposure Compensation")
        or ""
    )

    white_balance = (
        exif_map.get("WhiteBalance")
        or exif_map.get("White Balance")
        or ""
    )

    if white_balance == "0":
        white_balance = "Auto"
    elif white_balance == "1":
        white_balance = "Manual"

    return {
        "ISO": str(iso).strip(),
        "셔터스피드": str(shutter_speed).strip(),
        "초점거리": str(focal_length).strip(),
        "노출보정": str(exposure_bias).strip(),
        "화이트밸런스": str(white_balance).strip()
    }


# =========================================================
# 이미지 품질 / 밝기 / 중복
# =========================================================
def download_image_bytes(url: str):
    response = requests.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    return response.content


def check_image_quality(image_bytes: bytes):
    try:
        img = Image.open(BytesIO(image_bytes)).convert("RGB")

        if img.width < MIN_WIDTH or img.height < MIN_HEIGHT:
            return False, None

        colors = img.resize((128, 128)).getcolors(maxcolors=1000000)
        if colors is None or len(colors) < MIN_COLOR_COUNT:
            return False, None

        phash = imagehash.phash(img)
        return True, phash

    except Exception:
        return False, None


def get_brightness_info(image_bytes: bytes):
    try:
        img = Image.open(BytesIO(image_bytes)).convert("L")
        resized = img.resize((128, 128))
        pixels = list(resized.getdata())

        brightness_mean = sum(pixels) / len(pixels)

        if brightness_mean < 85:
            brightness_label = "dark"
        elif brightness_mean > 170:
            brightness_label = "bright"
        else:
            brightness_label = "normal"

        return brightness_label, round(brightness_mean, 2)

    except Exception:
        return "", ""


def is_similar_duplicate(new_hash, existing_hashes, threshold=PHASH_DUP_THRESHOLD):
    if new_hash is None:
        return False

    for old_hash in existing_hashes:
        if old_hash is None:
            continue
        if abs(new_hash - old_hash) <= threshold:
            return True

    return False


# =========================================================
# 셔터스피드 변환
# =========================================================
def shutter_to_seconds(shutter_value):
    try:
        value = str(shutter_value).strip()

        if value == "":
            return None

        # 예: "1/60"
        if "/" in value:
            return float(Fraction(value))

        # 예: "0.0167"
        return float(value)

    except Exception:
        return None


def seconds_to_shutter(seconds):
    if seconds is None:
        return ""

    try:
        seconds = float(seconds)

        if seconds <= 0:
            return ""

        if seconds >= 1:
            return f"{round(seconds, 2)}s"

        denominator = round(1 / seconds)
        return f"1/{denominator}"

    except Exception:
        return ""


# =========================================================
# 엑셀 저장
# =========================================================
def format_excel(excel_path: str):
    wb = load_workbook(excel_path)
    ws = wb.active

    center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = max(14, min(max_len + 2, 35))

    wb.save(excel_path)


def save_raw_excel(records: list, excel_path: str):
    df = pd.DataFrame(records)
    df.to_excel(excel_path, index=False)
    format_excel(excel_path)


def save_db_ready_excel(records: list, excel_path: str):
    if not records:
        return

    df = pd.DataFrame(records)

    if df.empty:
        return

    # ISO 숫자화
    df["ISO_숫자"] = pd.to_numeric(df["ISO"], errors="coerce")

    # 셔터스피드 초 단위 숫자화
    df["셔터스피드_초"] = df["셔터스피드"].apply(shutter_to_seconds)

    # 추천값 산출에 필요한 값만 사용
    df = df.dropna(subset=["ISO_숫자", "셔터스피드_초"])
    df = df[df["brightness"].isin(["dark", "normal", "bright"])]

    if df.empty:
        return

    summary = (
        df.groupby(["category", "brightness"])
        .agg(
            recommended_iso=("ISO_숫자", "median"),
            recommended_shutter_seconds=("셔터스피드_초", "median"),
            sample_count=("name", "count")
        )
        .reset_index()
    )

    summary["recommended_iso"] = summary["recommended_iso"].round().astype(int)
    summary["recommended_shutter"] = summary["recommended_shutter_seconds"].apply(seconds_to_shutter)

    summary = summary[
        [
            "category",
            "brightness",
            "recommended_iso",
            "recommended_shutter",
            "recommended_shutter_seconds",
            "sample_count"
        ]
    ]

    summary.to_excel(excel_path, index=False)
    format_excel(excel_path)


def save_all_db_ready_excel(all_records: list):
    excel_path = os.path.join(BASE_DIR, "all_categories_db_ready.xlsx")
    save_db_ready_excel(all_records, excel_path)
    print(f"전체 DB용 엑셀 저장 완료: {excel_path}")


# =========================================================
# 카테고리 하나 수집
# =========================================================
def collect_category(category_name: str):
    config = CATEGORY_CONFIG[category_name]
    category_dir = os.path.join(BASE_DIR, category_name)
    os.makedirs(category_dir, exist_ok=True)

    records = []
    saved_photo_ids = set()
    saved_hashes = []
    owner_count = {}

    current_count = 1

    print(f"\n{'=' * 60}")
    print(f"{category_name} 수집 시작")
    print(f"{'=' * 60}")

    for query in config["queries"]:
        if current_count > TARGET_COUNT:
            break

        print(f"\n검색어: {query}")
        no_progress_pages = 0

        for page in range(1, MAX_PAGES_PER_QUERY + 1):
            if current_count > TARGET_COUNT:
                break

            print(f"   - page {page} 검색 중...")
            before_count = current_count

            try:
                photos = search_photos(query, page)
            except Exception as e:
                print(f"   ! 검색 실패: {e}")
                break

            if not photos:
                print("   ! 검색 결과 없음 -> 다음 검색어")
                break

            for photo in photos:
                if current_count > TARGET_COUNT:
                    break

                photo_id = str(photo.get("id", "")).strip()
                owner_id = str(photo.get("owner", "")).strip()

                if not photo_id or photo_id in saved_photo_ids:
                    continue

                if owner_id and owner_count.get(owner_id, 0) >= MAX_IMAGES_PER_OWNER:
                    continue

                if not is_relevant_photo(
                    photo,
                    config["positive_keywords"],
                    config["negative_keywords"]
                ):
                    continue

                img_url = choose_image_url(photo)
                if not img_url:
                    continue

                # EXIF 검사
                exif_map = get_exif(photo_id)
                extracted = extract_required_fields(exif_map)

                # 핵심 정답값: ISO + 셔터스피드는 반드시 있어야 함
                if extracted["ISO"] == "" or extracted["셔터스피드"] == "":
                    continue

                # 이미지 다운로드
                try:
                    image_bytes = download_image_bytes(img_url)
                except Exception:
                    continue

                # 이미지 품질 검사 + pHash 생성
                is_good, phash = check_image_quality(image_bytes)
                if not is_good:
                    continue

                # 유사 이미지 제거
                if is_similar_duplicate(phash, saved_hashes, PHASH_DUP_THRESHOLD):
                    print("   ↪ 유사 이미지 스킵")
                    continue

                # 밝기 계산
                brightness_label, brightness_mean = get_brightness_info(image_bytes)

                file_name = f"{category_name}{current_count}.jpg"
                file_path = os.path.join(category_dir, file_name)

                try:
                    with open(file_path, "wb") as f:
                        f.write(image_bytes)
                except Exception:
                    continue

                record = {
                    "id": f"{category_name}_{current_count}",
                    "name": file_name,
                    "category": category_name,
                    "subcategory": query,
                    "brightness": brightness_label,
                    "brightness_mean": brightness_mean,
                    "ISO": extracted["ISO"],
                    "셔터스피드": extracted["셔터스피드"],
                    "셔터스피드_초": shutter_to_seconds(extracted["셔터스피드"]),
                    "초점거리": extracted["초점거리"],
                    "노출보정": extracted["노출보정"],
                    "화이트밸런스": extracted["화이트밸런스"],
                    "photo_id": photo_id,
                    "owner_id": owner_id,
                    "source": "flickr",
                    "image_path": file_path,
                    "title": photo.get("title", ""),
                    "tags": photo.get("tags", "")
                }

                records.append(record)
                saved_photo_ids.add(photo_id)
                saved_hashes.append(phash)

                if owner_id:
                    owner_count[owner_id] = owner_count.get(owner_id, 0) + 1

                print(f"   ✅ {category_name} {current_count}장 완료")
                current_count += 1

                time.sleep(0.05)

            if current_count == before_count:
                no_progress_pages += 1
                print(f"   진행 없음 ({no_progress_pages}/{NO_PROGRESS_LIMIT})")
            else:
                no_progress_pages = 0

                raw_excel_path = os.path.join(category_dir, f"{category_name}.xlsx")
                db_excel_path = os.path.join(category_dir, f"{category_name}_db_ready.xlsx")

                save_raw_excel(records, raw_excel_path)
                save_db_ready_excel(records, db_excel_path)

            if no_progress_pages >= NO_PROGRESS_LIMIT:
                print("   ➜ 이 검색어는 건너뜀")
                break

    raw_excel_path = os.path.join(category_dir, f"{category_name}.xlsx")
    db_excel_path = os.path.join(category_dir, f"{category_name}_db_ready.xlsx")

    save_raw_excel(records, raw_excel_path)
    save_db_ready_excel(records, db_excel_path)

    print(f"\n{category_name} 완료: 총 {len(records)}장")
    print(f"원본 엑셀: {raw_excel_path}")
    print(f"DB용 엑셀: {db_excel_path}")

    return records


def collect_all_categories():
    all_records = []

    for category_name in ["food", "landscape", "night", "portrait", "contrast"]:
        records = collect_category(category_name)
        all_records.extend(records)

    save_all_db_ready_excel(all_records)

    raw_all_path = os.path.join(BASE_DIR, "all_categories_raw.xlsx")
    save_raw_excel(all_records, raw_all_path)
    print(f"전체 원본 엑셀 저장 완료: {raw_all_path}")


if __name__ == "__main__":
    collect_category("contrast")
